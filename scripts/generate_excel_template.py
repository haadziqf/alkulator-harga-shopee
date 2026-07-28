import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Setup
    # ----------------------------------------------------
    header_fill = PatternFill(start_color="1E1512", end_color="1E1512", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    accent_fill = PatternFill(start_color="EE4D2D", end_color="EE4D2D", fill_type="solid")
    accent_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    highlight_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    highlight_font = Font(name="Calibri", size=12, bold=True, color="78350F")

    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=16, bold=True, color="1E1512")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="6E5E57")

    thin_border = Border(
        left=Side(style='thin', color='EAD8D0'),
        right=Side(style='thin', color='EAD8D0'),
        top=Side(style='thin', color='EAD8D0'),
        bottom=Side(style='thin', color='EAD8D0')
    )

    currency_format = 'Rp #,##0'
    percent_format = '0.0%'

    # ----------------------------------------------------
    # Sheet 1: Dashboard Laba Rugi
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "📊 Dashboard Laba Rugi"

    ws_dash['A1'] = "LAPORAN LABA RUGI PENJUALAN MARKETPLACE 2026"
    ws_dash['A1'].font = title_font
    ws_dash['A2'] = "Template Otomatis Pembukuan Toko Shopee, Tokopedia & TikTok Shop"
    ws_dash['A2'].font = subtitle_font

    headers_dash = ["Komponen Finansial", "Nilai (Rp)", "Persentase (%)", "Keterangan"]
    for col_num, h in enumerate(headers_dash, 1):
        cell = ws_dash.cell(row=4, column=col_num)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    financial_items = [
        ("Total Omzet Penjualan Kotor (Gross Sales)", "='🛒 Rekap Penjualan'!G20", "=B5/B5", "Total nilai transaksi penjualan sebelum potongan"),
        ("Total Potongan Komisi Marketplace", "='🛒 Rekap Penjualan'!H20", "=B6/B5", "Potongan admin, gratis ongkir & promo marketplace"),
        ("Total Uang Bersih Diterima (Net Revenue)", "=B5-B6", "=B7/B5", "Saldo bersih masuk ke rekening/saldo penjual"),
        ("Total HPP / Modal Produk", "='🛒 Rekap Penjualan'!J20", "=B8/B5", "Total akumulasi modal beli/produksi barang"),
        ("Total Biaya Packing", "='🛒 Rekap Penjualan'!K20", "=B9/B5", "Total plastik polymailer, kardus, bubble wrap"),
        ("Total Biaya Operasional & Iklan (Ads)", "='💸 Pengeluaran'!C20", "=B10/B5", "Pengeluaran iklan marketplace, gaji, software"),
        ("TOTAL LABA BERSIH (NET PROFIT)", "=B7-B8-B9-B10", "=B11/B5", "Keuntungan bersih akhir yang didapatkan toko"),
    ]

    for idx, (item, val_formula, pct_formula, note) in enumerate(financial_items, start=5):
        ws_dash.cell(row=idx, column=1, value=item)
        
        c_val = ws_dash.cell(row=idx, column=2, value=val_formula)
        c_val.number_format = currency_format
        
        c_pct = ws_dash.cell(row=idx, column=3, value=pct_formula)
        c_pct.number_format = percent_format
        
        ws_dash.cell(row=idx, column=4, value=note)

        if idx == 11: # Highlight LABA BERSIH
            for col in range(1, 5):
                cell = ws_dash.cell(row=idx, column=col)
                cell.fill = highlight_fill
                cell.font = highlight_font

    # ----------------------------------------------------
    # Sheet 2: Rekap Penjualan Transaksi
    # ----------------------------------------------------
    ws_sales = wb.create_sheet("🛒 Rekap Penjualan")
    
    headers_sales = [
        "Tanggal", "No. Pesanan", "Marketplace", "Nama Produk", 
        "Qty", "Harga Jual (Rp)", "Total Omzet (Rp)", "Komisi Marketplace (Rp)", 
        "Saldo Diterima (Rp)", "HPP Modal (Rp)", "Biaya Packing (Rp)", "Profit Bersih (Rp)"
    ]

    for col_num, h in enumerate(headers_sales, 1):
        cell = ws_sales.cell(row=1, column=col_num)
        cell.value = h
        cell.fill = accent_fill
        cell.font = accent_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_sales = [
        ("2026-07-01", "260701SHP001", "Shopee", "Kemeja Oversize Linen", 1, 150000, "=E2*F2", 21000, "=G2-H2", 85000, 3000, "=I2-J2-K2"),
        ("2026-07-01", "260701TKP002", "Tokopedia", "Celana Chino Slimfit", 2, 180000, "=E3*F3", 32400, "=G3-H3", 200000, 5000, "=I3-J3-K3"),
        ("2026-07-02", "260702TIK003", "TikTok Shop", "Kaos Polos Cotton Combed", 3, 65000, "=E4*F4", 22425, "=G4-H4", 105000, 4500, "=I4-J4-K4"),
        ("2026-07-02", "260702SHP004", "Shopee", "Jaket Varsity Vintage", 1, 280000, "=E5*F5", 39200, "=G5-H5", 160000, 4000, "=I5-J5-K5"),
    ]

    for row_idx, row_data in enumerate(sample_sales, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_sales.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in [6, 7, 8, 9, 10, 11, 12]:
                cell.number_format = currency_format

    # Add Total Summary Row at Row 20
    ws_sales.cell(row=20, column=1, value="TOTAL SUMMARY").font = bold_font
    for col_idx in range(5, 13):
        col_letter = get_column_letter(col_idx)
        cell = ws_sales.cell(row=20, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}19)")
        cell.font = bold_font
        if col_idx >= 6:
            cell.number_format = currency_format

    # ----------------------------------------------------
    # Sheet 3: Database Stok Barang & HPP
    # ----------------------------------------------------
    ws_stock = wb.create_sheet("📦 Database Stok & HPP")

    headers_stock = [
        "Kode SKU", "Nama Produk", "Kategori", "HPP Modal (Rp)", 
        "Biaya Packing (Rp)", "Harga Jual Disarankan (Rp)", "Stok Awal", "Terjual", "Sisa Stok", "Nilai Aset Stok (Rp)"
    ]

    for col_num, h in enumerate(headers_stock, 1):
        cell = ws_stock.cell(row=1, column=col_num)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_stock = [
        ("SKU-FSH-001", "Kemeja Oversize Linen", "Fashion", 85000, 3000, 150000, 100, 15, "=G2-H2", "=I2*D2"),
        ("SKU-FSH-002", "Celana Chino Slimfit", "Fashion", 100000, 2500, 180000, 80, 22, "=G3-H3", "=I3*D3"),
        ("SKU-FSH-003", "Kaos Polos Cotton Combed", "Fashion", 35000, 1500, 65000, 200, 45, "=G4-H4", "=I4*D4"),
        ("SKU-FSH-004", "Jaket Varsity Vintage", "Fashion", 160000, 4000, 280000, 50, 8, "=G5-H5", "=I5*D5"),
    ]

    for row_idx, row_data in enumerate(sample_stock, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_stock.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in [4, 5, 6, 10]:
                cell.number_format = currency_format

    # ----------------------------------------------------
    # Sheet 4: Rekap Pengeluaran Operasional & Iklan
    # ----------------------------------------------------
    ws_exp = wb.create_sheet("💸 Pengeluaran")

    headers_exp = ["Tanggal", "Kategori Pengeluaran", "Nominal Pengeluaran (Rp)", "Keterangan / Catatan"]

    for col_num, h in enumerate(headers_exp, 1):
        cell = ws_exp.cell(row=1, column=col_num)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_exp = [
        ("2026-07-01", "Iklan Shopee Ads", 150000, "Top-up saldo iklan Shopee"),
        ("2026-07-02", "Iklan TikTok Shop (GMV Max)", 200000, "Kempen promosi tanggal kembar"),
        ("2026-07-03", "Bahan Packing (Plastik Polymailer)", 85000, "Beli 100 pcs plastik ukuran L"),
        ("2026-07-05", "Gaji Admin Packing", 500000, "Uang harian tim packing pesanan"),
    ]

    for row_idx, row_data in enumerate(sample_exp, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_exp.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 3:
                cell.number_format = currency_format

    # Total Summary for Expenses at Row 20
    ws_exp.cell(row=20, column=1, value="TOTAL PENGELUARAN").font = bold_font
    cell_tot = ws_exp.cell(row=20, column=3, value="=SUM(C2:C19)")
    cell_tot.font = bold_font
    cell_tot.number_format = currency_format

    # Auto-adjust Column Widths for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save("public/Template_Pembukuan_Toko_Marketplace_2026.xlsx")
    print("✅ Excel template successfully generated at public/Template_Pembukuan_Toko_Marketplace_2026.xlsx!")

if __name__ == "__main__":
    create_template()
